import unittest
import openpyxl
import glob
import importlib
import timeout_decorator
import os
import time
import signal
import logging

# Define the name of the Excel file
RESULT_FILE = "summary.xlsx"
# Open the existing Excel file
wb = openpyxl.load_workbook(RESULT_FILE)
# Get the sheet you want to write to
SHEET_NAME = 'Result'
# Define the folder containing the modules
module_folder = 'solutions'
# Timeout duration in seconds for each function
TIMEOUT_DURATION = 1
# Timeout duration in seconds for each signal
TIMEOUT_SIGNAL = 3
# Initialize logging
LOGGING_FILE = 'result.log'

def init_logging():
    global LOGGING_FILE, logging
    # Delete result_log if it exists
    if os.path.exists('result_log'):
        os.remove('result_log')
        logging.info("Deleted existing result_log file")
    # Configure logging
    logging.basicConfig(filename=LOGGING_FILE, level=logging.INFO, 
                        format='%(asctime)s - %(levelname)s - %(message)s')

def init_sheet():
    global wb

    sheet = wb[SHEET_NAME]
    wb.remove(sheet)

    # create an empty sheet [sheet_name] using old index
    wb.create_sheet(SHEET_NAME)

    # Write the data to the sheet
    sheet = wb[SHEET_NAME]
    sheet.cell(row=1, column=1).value = "Student File"
    sheet.cell(row=1, column=2).value = "Total Test"
    sheet.cell(row=1, column=3).value = "Pass Test"
    sheet.cell(row=1, column=4).value = "Score"

class AutoMationTest(unittest.TestCase):
    currentResult = None # holds last result object passed to run method

    @classmethod
    def setResult(cls, amount, errors, failures, skipped):
        cls.amount, cls.errors, cls.failures, cls.skipped = \
            amount, errors, failures, skipped

    def tearDown(self):
        amount = self.currentResult.testsRun
        errors = self.currentResult.errors
        failures = self.currentResult.failures
        skipped = self.currentResult.skipped
        self.setResult(amount, errors, failures, skipped)

    @classmethod
    def tearDownClass(cls):
        logging.info(f"\ntests run: {cls.amount}")
        logging.info(f"errors: {len(cls.errors)}")
        logging.info(f"failures: {len(cls.failures)}")
        logging.info(f"success: {cls.amount - len(cls.errors) - len(cls.failures)}")
        logging.info(f"skipped: {len(cls.skipped)}")

    def run(self, result=None):
        self.currentResult = result # remember result for use in tearDown
        unittest.TestCase.run(self, result) # call superclass run method

class AutomationTestModule(AutoMationTest):
    def __init__(self, testName, module):
        super(AutomationTestModule, self).__init__(testName)
        self.module = module

    def writeToExcel(self):
        sheet = wb[SHEET_NAME]
        last_row = sheet.max_row
        sheet.cell(row=last_row+1, column=1).value = str(self.module.__name__.split('.')[-1])
        sheet.cell(row=last_row+1, column=2).value = str(self.currentResult.testsRun)
        sheet.cell(row=last_row+1, column=3).value = str(self.currentResult.testsRun - len(self.currentResult.errors) - len(self.currentResult.failures))
        sheet.cell(row=last_row+1, column=4).value = str((self.currentResult.testsRun - len(self.currentResult.errors) - len(self.currentResult.failures))*10/self.currentResult.testsRun)

class TestCase(AutomationTestModule):
    def __init__(self, testName, module):
        AutomationTestModule.__init__(self, testName, module)

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testXuatChuoiNhiPhan1(self):
        self.assertEqual(self.module.xuatChuoiNhiPhan(-10), '-1')

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testXuatChuoiNhiPhan2(self):
        self.assertEqual(self.module.xuatChuoiNhiPhan(5), '101')

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testXuatChuoiNhiPhan3(self):
        self.assertEqual(self.module.xuatChuoiNhiPhan(100), '1100100')

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testChuoiPalindrome1(self):
        self.assertEqual(self.module.laPalindrome(-100), False)
    
    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testChuoiPalindrome2(self):
        self.assertEqual(self.module.laPalindrome(100000), False)

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testChuoiPalindrome3(self):
        self.assertEqual(self.module.laPalindrome(121121), True)

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testSoNguyenTo1(self):
        self.assertEqual(self.module.laSoNguyenTo(-100), False)

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testSoNguyenTo2(self):
        self.assertEqual(self.module.laSoNguyenTo(19), True)

    @timeout_decorator.timeout(TIMEOUT_DURATION)
    def testTongSoNguyenTo(self):
        self.assertEqual(self.module.tongSoNguyenTo(100), 1060)
        
class TimeoutException(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutException

def run_tests(module_file: str, module_folder: str, result_file: str):
    try:
        module_name = module_file.split('/')[-1].split('.')[0]
        logging.info(f"Importing module: {module_name}")
        
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(TIMEOUT_SIGNAL)
        
        try:
            start_time = time.time()  # Start timer
            module = importlib.import_module('.' + module_name, package=module_folder)
            end_time = time.time()  # End timer
            signal.alarm(0)  # Disable the alarm
            
            duration = end_time - start_time
            logging.info(f"Module import took {duration:.4f} seconds")
        except TimeoutException:
            logging.warning(f"Module {module_name} import timed out after {TIMEOUT_SIGNAL} seconds")
            module = None  # Ignore the module if it takes too long
        
        if module:
            suite = unittest.TestSuite()

            suite.addTest(TestCase("testXuatChuoiNhiPhan1", module))
            suite.addTest(TestCase("testXuatChuoiNhiPhan2", module))
            suite.addTest(TestCase("testXuatChuoiNhiPhan3", module))
            suite.addTest(TestCase("testChuoiPalindrome1", module))
            suite.addTest(TestCase("testChuoiPalindrome2", module))
            suite.addTest(TestCase("testChuoiPalindrome3", module))
            suite.addTest(TestCase("testSoNguyenTo1", module))
            suite.addTest(TestCase("testSoNguyenTo2", module))
            suite.addTest(TestCase("testTongSoNguyenTo", module))

            suite.addTest(TestCase("writeToExcel", module))

            runner = unittest.TextTestRunner()
            runner.run(suite)
    except Exception as e:
        logging.error(f"{module_name} error!!! {e}")

    # Save the changes to the Excel file
    wb.save(RESULT_FILE)
    
if __name__ == '__main__':
    init_logging()
    init_sheet()
    # Get a list of all Python files in the folder
    module_files = glob.glob(module_folder + '/*.py')
    for module_file in module_files:
        run_tests(module_file, module_folder, RESULT_FILE)
