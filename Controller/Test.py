import unittest
import openpyxl
import glob
import importlib
import Database.Store as Store
import os
import time
import signal
import logging

from Environment.const import TEST_LOGGING_FILE, TEST_RESULT_FILE, TEST_SHEET_NAME, TEST_STUDENT_EXAM_FOL, TEST_TIMEOUT_DURATION, TEST_TIMEOUT_SIGNAL

class AutoMationTest(unittest.TestCase):
    currentResult = None # holds last result object passed to run method

    @classmethod
    def setResult(cls, amount, errors, failures, skipped):
        cls.amount, cls.errors, cls.failures, cls.skipped = \
            amount, errors, failures, skipped

    def tearDown(self):
        amount = self.currentResult.testsRun
        errors = self.currentResult.errors
        failures = self.currentResult.failures
        skipped = self.currentResult.skipped
        self.setResult(amount, errors, failures, skipped)

    @classmethod
    def tearDownClass(cls):
        logging.info(f"\ntests run: {cls.amount}")
        logging.info(f"errors: {len(cls.errors)}")
        logging.info(f"failures: {len(cls.failures)}")
        logging.info(f"success: {cls.amount - len(cls.errors) - len(cls.failures)}")
        logging.info(f"skipped: {len(cls.skipped)}")

    def run(self, result=None):
        self.currentResult = result # remember result for use in tearDown
        unittest.TestCase.run(self, result) # call superclass run method

class AutomationTestModule(AutoMationTest):
    def __init__(self, testName, module):
        super(AutomationTestModule, self).__init__(testName)
        self.module = module

    def writeToExcel(self):
        sheet = self.wb[TEST_SHEET_NAME]
        last_row = sheet.max_row
        sheet.cell(row=last_row+1, column=1).value = str(self.module.__name__.split('.')[-1])
        sheet.cell(row=last_row+1, column=2).value = str(self.currentResult.testsRun)
        sheet.cell(row=last_row+1, column=3).value = str(self.currentResult.testsRun - len(self.currentResult.errors) - len(self.currentResult.failures))
        sheet.cell(row=last_row+1, column=4).value = str((self.currentResult.testsRun - len(self.currentResult.errors) - len(self.currentResult.failures))*10/self.currentResult.testsRun)

class TestCase(AutomationTestModule):
    test_case_from_db = None
    test_cases = None
    
    def __init__(self, testName, module):
        AutomationTestModule.__init__(self, testName, module)
        store = Store()
        self.test_case_from_db = store.read_data_from_db()
        
    def runTest(self):
        for function_name, input, output in self.test_case_from_db:
            self.test_cases.append((f'test{function_name}', self.module.function_name, input, output))
        # test_cases = [
        #     ("testXuatChuoiNhiPhan1", self.module.xuatChuoiNhiPhan, -10, '-1'),
        #     ("testXuatChuoiNhiPhan2", self.module.xuatChuoiNhiPhan, 5, '101'),
        #     ("testXuatChuoiNhiPhan3", self.module.xuatChuoiNhiPhan, 100, '1100100'),
        #     ("testChuoiPalindrome1", self.module.laPalindrome, -100, False),
        #     ("testChuoiPalindrome2", self.module.laPalindrome, 100000, False),
        #     ("testChuoiPalindrome3", self.module.laPalindrome, 121121, True),
        #     ("testSoNguyenTo1", self.module.laSoNguyenTo, -100, False),
        #     ("testSoNguyenTo2", self.module.laSoNguyenTo, 19, True),
        #     ("testTongSoNguyenTo", self.module.tongSoNguyenTo, 100, 1060)
        # ]

        for test_name, func, input_val, expected in self.test_cases:
            with self.subTest(test_name=test_name):
                self.assertEqual(func(input_val), expected)

class TimeoutException(Exception):
    pass

class Test():
    wb = None
    
    def __init__(self):
        self.init_logging()
        self.init_sheet()
        self.wb = openpyxl.load_workbook(TEST_RESULT_FILE)
        
    def get_user_input_files(self):
        module_files = glob.glob(TEST_STUDENT_EXAM_FOL + '/*.py')
        return module_files
    
    def run_tests_on_files(self):
        module_files = self.get_user_input_files()
        for module_file in module_files:
            self.run_tests(module_file, TEST_STUDENT_EXAM_FOL, TEST_RESULT_FILE)

    def init_logging():
        if os.path.exists('result_log'):
            os.remove('result_log')
            logging.info("Deleted existing result_log file")
        logging.basicConfig(filename=TEST_LOGGING_FILE, level=logging.INFO, 
                            format='%(asctime)s - %(levelname)s - %(message)s')

    def init_sheet(self):
        sheet = self.wb[TEST_SHEET_NAME]

        self.wb.remove(sheet)
        self.wb.create_sheet(TEST_SHEET_NAME)

        sheet = self.wb[TEST_SHEET_NAME]
        sheet.cell(row=1, column=1).value = "Student File"
        sheet.cell(row=1, column=2).value = "Total Test"
        sheet.cell(row=1, column=3).value = "Pass Test"
        sheet.cell(row=1, column=4).value = "Score"
        
    def timeout_handler(self, signum, frame):
        raise TimeoutException

    def run_tests(self, module_file: str, TEST_STUDENT_EXAM_FOL: str, TEST_RESULT_FILE: str):
        try:
            module_name = module_file.split('/')[-1].split('.')[0]
            logging.info(f"Importing module: {module_name}")
            
            signal.signal(signal.SIGALRM, self.timeout_handler)
            signal.alarm(TEST_TIMEOUT_SIGNAL)
            
            try:
                start_time = time.time()  # Start timer
                module = importlib.import_module('.' + module_name, package=TEST_STUDENT_EXAM_FOL)
                end_time = time.time()  # End timer
                signal.alarm(0)  # Disable the alarm
                
                duration = end_time - start_time
                logging.info(f"Module import took {duration:.4f} seconds")
            except TimeoutException:
                logging.warning(f"Module {module_name} import timed out after {TEST_TIMEOUT_SIGNAL} seconds")
                module = None  # Ignore the module if it takes too long
            
            if module:
                suite = unittest.TestSuite()

                # suite.addTest(TestCase("testXuatChuoiNhiPhan1", module))
                # suite.addTest(TestCase("testXuatChuoiNhiPhan2", module))
                # suite.addTest(TestCase("testXuatChuoiNhiPhan3", module))
                # suite.addTest(TestCase("testChuoiPalindrome1", module))
                # suite.addTest(TestCase("testChuoiPalindrome2", module))
                # suite.addTest(TestCase("testChuoiPalindrome3", module))
                # suite.addTest(TestCase("testSoNguyenTo1", module))
                # suite.addTest(TestCase("testSoNguyenTo2", module))
                # suite.addTest(TestCase("testTongSoNguyenTo", module))

                suite.addTest(TestCase("writeToExcel", module))

                runner = unittest.TextTestRunner()
                runner.run(suite)
        except Exception as e:
            logging.error(f"{module_name} error!!! {e}")

        # Save the changes to the Excel file
        self.wb.save(TEST_RESULT_FILE)