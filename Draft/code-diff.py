# importing SequenceMatcher of difflib module
from difflib import SequenceMatcher
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import column_index_from_string

import os
import openpyxl
import chardet
import pandas as pd


# Open the existing Excel file
resultFileName = "summary.xlsx"
wb = openpyxl.load_workbook(resultFileName)
# Get the sheet you want to write to
sheetName = 'Plagiarized'
# Sort by column A in ascending order

def initSheet():
    global wb
    sheet = wb[sheetName]
    wb.remove(sheet)

    # create an empty sheet [sheet_name] using old index
    wb.create_sheet(sheetName)
    sheet = wb[sheetName]
    sheet.cell(row=1, column=1).value = "File 1"
    sheet.cell(row=1, column=2).value = "File 2"
    sheet.cell(row=1, column=3).value = "Plagiarized %"
    sheet.cell(row=1, column=4).value = "Diff"

dirPath = os.path.join(os.getcwd(), "./solutions")
print(dirPath)

def writeToExcel(fileName1, fileName2, plagiarized, diffCode):
    sheet = wb[sheetName]
    last_row = sheet.max_row
    sheet.cell(row=last_row+1, column=1).value = str(fileName1)
    sheet.cell(row=last_row+1, column=2).value = str(fileName2)
    sheet.cell(row=last_row+1, column=3).value = str(plagiarized)
    sheet.cell(row=last_row+1, column=4).value = ILLEGAL_CHARACTERS_RE.sub(r'',diffCode)

def exportDiff(seqm):
    output= []
    for opcode, a0, a1, b0, b1 in seqm.get_opcodes():
        if opcode == 'equal':
            output.append(seqm.a[a0:a1])
        elif opcode == 'insert':
            output.append("<ins>" + seqm.b[b0:b1] + "</ins>")
        elif opcode == 'delete':
            output.append("<del>" + seqm.a[a0:a1] + "</del>")
        elif opcode == 'replace':
            output.append("<mark>" + seqm.a[a0:a1] + " -> " + seqm.b[b0:b1] + "</mark>")
        else:
            raise RuntimeError("unexpected opcode")
    return ''.join(output)

def checkPlagiarized(filePath1: str, filePath2: str):
    file1Encoding = chardet.detect(open(filePath1, 'rb').read())['encoding']
    file2Encoding = chardet.detect(open(filePath2, 'rb').read())['encoding']
    with open(filePath1, encoding=file1Encoding) as first_file, open(filePath2, encoding=file2Encoding) as second_file:
        # Reading Both Text Files
        file1 = first_file.read()
        file2 = second_file.read()
        
        # Comparing Both Text Files
        sm = SequenceMatcher(None, file1,
                            file2)
        
        # converting decimal output in integer
        plagiarizedPercent = int(sm.ratio()*100)
        diffCode = exportDiff(sm)
        writeToExcel(os.path.basename(filePath1), os.path.basename(filePath2), plagiarizedPercent, diffCode)

def traversePair(dirPath, files):
    for i in range(0, len(files)):
        for j in range(i + 1, len(files)):
            filePath1 = os.path.join(dirPath, files[i])
            filePath2 = os.path.join(dirPath, files[j])
            checkPlagiarized(filePath1, filePath2)

if __name__ == "__main__":
    initSheet()
    files = [f for f in os.listdir(dirPath) if os.path.isfile(os.path.join(dirPath, f))]
    traversePair(dirPath, files)
    wb.save(resultFileName)